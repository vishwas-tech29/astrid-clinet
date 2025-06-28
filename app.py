from flask import Flask, request, jsonify
from datetime import datetime
import csv
import os
from openpyxl import Workbook, load_workbook

app = Flask(__name__)

# File paths
CSV_FILE = 'appointments.csv'
EXCEL_FILE = 'appointments.xlsx'

@app.route('/submit-appointment', methods=['POST'])
def submit_appointment():
    try:
        data = request.get_json()
        
        # Validate required fields
        required_fields = ['name', 'email', 'phone', 'service', 'date']
        for field in required_fields:
            if not data.get(field):
                return jsonify({'success': False, 'message': f'Missing required field: {field}'}), 400
        
        # Save to CSV
        file_exists = os.path.isfile(CSV_FILE)
        with open(CSV_FILE, 'a', newline='') as csvfile:
            fieldnames = ['name', 'email', 'phone', 'service', 'date', 'message', 'timestamp']
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
            
            if not file_exists:
                writer.writeheader()
            
            writer.writerow(data)
        
        # Save to Excel
        if os.path.exists(EXCEL_FILE):
            wb = load_workbook(EXCEL_FILE)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.append(['Name', 'Email', 'Phone', 'Service', 'Date', 'Message', 'Timestamp'])
        
        ws.append([
            data['name'],
            data['email'],
            data['phone'],
            data['service'],
            data['date'],
            data.get('message', ''),
            data['timestamp']
        ])
        wb.save(EXCEL_FILE)
        
        return jsonify({'success': True, 'message': 'Appointment booked successfully!'})
    
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

if __name__ == '__main__':
    app.run(debug=True)